# app.py
# -*- coding: utf-8 -*-

import pyodbc
from barcode import Code128
from barcode.writer import ImageWriter
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, make_response, flash, Response
from werkzeug.security import check_password_hash, generate_password_hash
from functools import wraps
import os
import io
import csv
from datetime import datetime, timedelta
from weasyprint import HTML, CSS
import requests 
import json
import traceback

# Importações para QRCode, E-mail e EXCEL
import qrcode
import smtplib
import ssl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import base64
from io import BytesIO

# --- Configuração da Aplicação Flask ---
app = Flask(__name__)
app.secret_key = os.urandom(24) 

# --- FILTRO PERSONALIZADO PARA FORMATAR MOEDA NO PADRÃO BR ---
def format_currency(value):
    """Formata um número como moeda no padrão R$ 1.234,56"""
    try:
        val = float(value)
        formatted_val = f'{val:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
        return f"R$ {formatted_val}"
    except (ValueError, TypeError):
        return "R$ 0,00"
app.jinja_env.filters['currency'] = format_currency

# --- CONFIGURAÇÕES DE E-MAIL ---
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
EMAIL_SENDER = "reunidassocial@gmail.com"
EMAIL_PASSWORD = "udpymgqfzlujgrpl"

# --- CHAVE SECRETA PARA A API DE SINCRONIZAÇÃO ---
SYNC_SECRET_TOKEN = "Rp@T3ch#50" 

# --- Configurações de Conexão com os Bancos de Dados ---
DB_SERVER = '172.16.1.223'
DB_USERNAME = 'sa'
DB_PASSWORD = 'Rp@T3ch#50'

DB_BI = 'p12_BI'
DB_INVENTARIO = 'inventario'

DB_CONNECTION_STRING = (
    f'DRIVER={{ODBC Driver 17 for SQL Server}};'
    f'SERVER={DB_SERVER};'
    f'DATABASE={DB_INVENTARIO};'
    f'UID={DB_USERNAME};'
    f'PWD={DB_PASSWORD};'
    f'TrustServerCertificate=yes;'
)

# --- GRUPOS DE ALMOXARIFADO PARA CONTAGEM CONSOLIDADA ---
ALMOXARIFADOS_GRUPO = {
    'Grãos': ['0201', '0203', '0301'],
    'Cana': ['0801', '0802'],
    'Armazém': ['0701']
}

# --- LISTA DE MENUS DISPONÍVEIS NO SISTEMA ---
AVAILABLE_MENUS = {
    'dashboard': 'Contagem Detalhada',
    'contagem_consolidada_page': 'Contagem Consolidada',
    'contagem_planejada': 'Contagem Planejada',
    'lista_recontagem': 'Recontagem',
    'gerar_etiquetas': 'Gerar Etiquetas',
    'gerenciar_inventarios': 'Inventários',
    'selecionar_progress_dashboard': 'Progresso',
    'selecionar_inventario_analitico': 'Dashboard Analítico',
    'analise_estoque_parado': 'Estoque Parado',
    'admin': 'Admin (Usuários)',
    'manage_roles': 'Admin (Perfis)',
    'view_logs': 'Logs da API',
    # --- NOVAS PERMISSÕES ADICIONADAS ---
    'edit_user': 'Admin (Editar Usuário)',
    'reset_password': 'Admin (Resetar Senha)',
    'delete_user': 'Admin (Excluir Usuário)'
}

# --- FUNÇÃO DE LOGGING NO BANCO DE DADOS ---
def log_to_db(level, message, source='ReceiverAPI (223)'):
    """Grava uma mensagem de log na tabela sync_logs e retorna o ID do novo log."""
    new_log_id = None
    try:
        conn_log = get_db_connection(DB_INVENTARIO)
        if conn_log:
            cursor_log = conn_log.cursor()
            # Adiciona OUTPUT INSERTED.ID para obter o ID do novo registo
            sql = "INSERT INTO sync_logs (log_level, source, message) OUTPUT INSERTED.ID VALUES (?, ?, ?)"
            cursor_log.execute(sql, level, source, message)
            new_log_id = cursor_log.fetchone()[0]
            conn_log.commit()
            conn_log.close()
    except Exception as e:
        print(f"[{datetime.now()}] FALHA CRÍTICA AO GRAVAR LOG: {e}")
    return new_log_id

# --- Funções de Conexão ---
def get_db_connection(database_name):
    try:
        conn_str = (
            f'DRIVER={{ODBC Driver 17 for SQL Server}};'
            f'SERVER={DB_SERVER};'
            f'DATABASE={database_name};'
            f'UID={DB_USERNAME};'
            f'PWD={DB_PASSWORD};'
            f'TrustServerCertificate=yes;'
        )
        conn = pyodbc.connect(conn_str)
        return conn
    except pyodbc.Error as ex:
        log_to_db('ERROR', f"Erro de conexão com o banco '{database_name}': {ex}")
        return None

# --- Decorators ---
# --- SISTEMA DE AUTENTICAÇÃO E PERMISSÃO ATUALIZADO ---

@app.before_request
def load_user_permissions():
    """Carrega as permissões do usuário em cada requisição."""
    if 'user_id' in session:
        # Evita recarregar a cada request se as permissões já estiverem na sessão
        if 'permissions' not in session:
            # LÓGICA CORRIGIDA: Trata o Admin como um caso especial
            if session.get('role_name') == 'Admin':
                # Se for Admin, concede todas as permissões diretamente
                session['permissions'] = list(AVAILABLE_MENUS.keys())
            elif 'role_id' in session:
                # Para outros perfis, busca no banco de dados
                conn = get_db_connection(DB_INVENTARIO)
                cursor = conn.cursor()
                cursor.execute("SELECT menu_endpoint FROM role_permissions WHERE role_id = ?", session['role_id'])
                session['permissions'] = [row.menu_endpoint for row in cursor.fetchall()]
                conn.close()
            else:
                session['permissions'] = [] # Garante que a lista exista, mesmo que vazia

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash("Por favor, faça login para acessar esta página.", 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# NOVO DECORATOR: Para verificar se o usuário é Administrador
def admin_required(f):
  @wraps(f)
  def decorated_function(*args, **kwargs):
    if 'user_id' not in session:
      flash("Por favor, faça login para acessar esta página.", 'warning')
      return redirect(url_for('login'))
   
    # CORREÇÃO: Altera a verificação para usar o nome do perfil
    if session.get('role_name') != 'Admin':
      flash("Você não tem permissão para acessar esta página.", 'error')
      return redirect(url_for('dashboard'))
   
    return f(*args, **kwargs)
  return decorated_function

def role_required(*roles):
  def wrapper(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
      # CORREÇÃO: Use a chave de sessão 'role_name'
      if session.get('role_name') not in roles:
        flash('Acesso negado. Você não tem permissão para acessar esta página.', 'error')
        return redirect(url_for('dashboard'))
      return f(*args, **kwargs)
    return decorated_function
  return wrapper

# --- ROTA DE API PARA SINCRONIZAÇÃO ---
@app.route('/api/sync_balances', methods=['POST'])
def sync_balances_receiver():
    # Depois de confirmarmos que funciona, pode apagar estas duas linhas de print
    print(f"DEBUG: Payload recebido: {request.json}")
    
    token = request.headers.get('X-Sync-Token')
    if token != SYNC_SECRET_TOKEN:
        log_to_db('WARNING', f"Tentativa de acesso não autorizada à API de sync do IP: {request.remote_addr}")
        return jsonify({"error": "Não autorizado"}), 403

    items_to_update = request.json
    if not items_to_update:
        return jsonify({"message": "Nenhum item para atualizar."}), 200

    sample_codes = ', '.join([f"{item['code']}({item['filial']}-{item['local']})" for item in items_to_update[:3]])
    log_to_db('INFO', f"Recebido pedido de sincronização para {len(items_to_update)} itens. Amostra: [{sample_codes}...]")

    conn_dest = get_db_connection(DB_BI)
    if not conn_dest:
        log_to_db('ERROR', "Falha ao conectar ao banco de dados de destino (p12_BI) para sincronização.")
        return jsonify({"error": "Falha na conexão com o banco de dados"}), 500
    
    cursor_dest = conn_dest.cursor()
    
    try:
        old_quantities = {}
        for item in items_to_update:
            cursor_dest.execute(
                "SELECT B2_QATU FROM SB2010 WHERE B2_COD = ? AND B2_FILIAL = ? AND B2_LOCAL = ?",
                item['code'], item['filial'], item['local']
            )
            row = cursor_dest.fetchone()
            key = (item['code'], item['filial'], item['local'])
            old_quantities[key] = row.B2_QATU if row else 0

        update_query = "UPDATE dbo.SB2010 SET B2_QATU = ? WHERE B2_COD = ? AND B2_FILIAL = ? AND B2_LOCAL = ?"
        update_data = [(item['balance'], item['code'], item['filial'], item['local']) for item in items_to_update]
        cursor_dest.executemany(update_query, update_data)
        conn_dest.commit()
        
        log_message = f"Sincronização bem-sucedida: {len(update_data)} itens atualizados. Amostra: [{sample_codes}...]"
        log_id = log_to_db('INFO', log_message, source='ReceiverAPI (223)')

        if log_id:
            conn_log = get_db_connection(DB_INVENTARIO)
            if conn_log:
                cursor_log = conn_log.cursor()
                details_query = """
                INSERT INTO sync_log_details 
                (log_id, item_code, previous_quantity, quantity_sent, item_description, filial, local) 
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """
                
                # --- ALTERAÇÃO PRINCIPAL AQUI ---
                # Em vez de executemany, fazemos um loop e executamos um por um.
                for item in items_to_update:
                    key = (item['code'], item['filial'], item['local'])
                    params = (
                        log_id, 
                        item['code'], 
                        old_quantities.get(key, 0),
                        item['balance'], 
                        item.get('description', 'N/A'),
                        item['filial'],
                        item['local']
                    )
                    cursor_log.execute(details_query, params)
                # ---------------------------------
                
                conn_log.commit()
                conn_log.close()

        return jsonify({"message": f"{len(update_data)} itens sincronizados com sucesso."}), 200
    except pyodbc.Error as e:
        conn_dest.rollback()
        log_to_db('ERROR', f"Erro de banco de dados durante a sincronização: {e}", source='ReceiverAPI (223)')
        return jsonify({"error": "Erro ao atualizar os dados"}), 500
    finally:
        if conn_dest:
            conn_dest.close()


# --- NOVA ROTA PARA VISUALIZAR DETALHES DO LOG ---
@app.route('/admin/log_details/<int:log_id>')
@login_required
@role_required('Admin')
def view_log_details(log_id):
    conn = get_db_connection(DB_INVENTARIO)
    cursor = conn.cursor()
    
    cursor.execute("SELECT id, log_timestamp, source, message FROM sync_logs WHERE id = ?", log_id)
    log_entry = cursor.fetchone()
    
    cursor.execute("""
        SELECT item_code, quantity_sent, previous_quantity, item_description, filial, local 
        FROM sync_log_details 
        WHERE log_id = ? 
        ORDER BY item_code
    """, log_id)
    details = cursor.fetchall()
    
    conn.close()
    
    if not log_entry:
        flash('Log não encontrado.', 'error')
        return redirect(url_for('view_logs'))
        
    return render_template('log_details.html', log=log_entry, details=details, username=session.get('username'))

# --- NOVA ROTA PARA VISUALIZAR LOGS ---
@app.route('/admin/logs')
@login_required
@role_required('Admin')
def view_logs():
    # Pega todos os possíveis filtros da URL como texto
    search_query = request.args.get('search', '').strip()
    start_date_str = request.args.get('start_date', '').strip()
    end_date_str = request.args.get('end_date', '').strip()
    
    conn = get_db_connection(DB_INVENTARIO)
    cursor = conn.cursor()
    
    sql_query_base = "SELECT id, log_timestamp, log_level, source, message FROM sync_logs"
    where_clauses = []
    params = []
    
    # Adiciona o filtro de texto se existir
    if search_query:
        where_clauses.append("message LIKE ?")
        params.append(f"%{search_query}%")
    
    # --- LÓGICA DE CONVERSÃO DE DATA AQUI ---
    try:
        if start_date_str:
            # Converte o texto para um objeto datetime
            start_date_obj = datetime.strptime(start_date_str, '%Y-%m-%dT%H:%M')
            where_clauses.append("log_timestamp >= ?")
            params.append(start_date_obj) # Envia o objeto para o pyodbc
            
        if end_date_str:
            # Converte o texto para um objeto datetime
            end_date_obj = datetime.strptime(end_date_str, '%Y-%m-%dT%H:%M')
            where_clauses.append("log_timestamp <= ?")
            params.append(end_date_obj) # Envia o objeto para o pyodbc

    except ValueError:
        flash("Formato de data inválido. Por favor, use o seletor de data e hora.", 'error')
        # Em caso de erro, não aplica o filtro de data para evitar crash
        pass
    # --- FIM DA LÓGICA DE CONVERSÃO ---

    # Constrói a query final
    if where_clauses:
        final_query = f"{sql_query_base} WHERE {' AND '.join(where_clauses)} ORDER BY log_timestamp DESC"
    else:
        final_query = f"SELECT TOP 200 id, log_timestamp, log_level, source, message FROM sync_logs ORDER BY log_timestamp DESC"
    
    cursor.execute(final_query, *params)
    logs = cursor.fetchall()
    conn.close()
    
    # Passa os textos originais de volta para o template para preencher os campos
    return render_template('logs.html', 
                           logs=logs, 
                           search_query=search_query, 
                           start_date=start_date_str,
                           end_date=end_date_str,
                           username=session.get('username'))

@app.route('/')
def index():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        conn = get_db_connection(DB_INVENTARIO)
        if conn:
            cursor = conn.cursor()
            # Query atualizada para buscar dados da tabela roles
            cursor.execute("""
                SELECT u.id, u.username, u.password_hash, r.id as role_id, r.name as role_name
                FROM usuarios u
                JOIN roles r ON u.role_id = r.id
                WHERE u.username = ? AND u.D_E_L_E_T_ <> '*' AND r.D_E_L_E_T_ <> '*'
            """, username)
            user = cursor.fetchone()
            conn.close()
            if user and check_password_hash(user.password_hash, password):
                session['user_id'] = user.id
                session['username'] = user.username
                session['role_id'] = user.role_id
                session['role_name'] = user.role_name # Salva o nome do perfil
                return redirect(url_for('dashboard'))
        flash('Usuário ou senha inválidos.', 'error')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    sessao_id = request.args.get('sessao_id')
    item_code = request.args.get('item_code')

    conn_inventario = get_db_connection(DB_INVENTARIO)
    cursor_inventario = conn_inventario.cursor()
    cursor_inventario.execute("SELECT id, descricao FROM inventario_sessoes WHERE status = 'Aberto' AND D_E_L_E_T_ <> '*' ORDER BY data_abertura DESC")
    sessoes_abertas = cursor_inventario.fetchall()

    conn_inventario.close()

    return render_template('dashboard.html',
                           sessoes=sessoes_abertas,
                           sessao_id=sessao_id,
                           item_code=item_code,
                           # CORREÇÃO: Passa o dicionário ALMOXARIFADOS_GRUPO para o template
                           ALMOXARIFADOS_GRUPO=ALMOXARIFADOS_GRUPO,
                           username=session.get('username'))

@app.route('/admin', methods=['GET', 'POST'])
@admin_required
def admin():
    conn = None
    try:
        conn = pyodbc.connect(DB_CONNECTION_STRING)
        cursor = conn.cursor()

        if request.method == 'POST':
            username = request.form['username']
            password = request.form['password']
            filial = request.form['filial']
            role_id = request.form['role_id']
            pode_ver_saldo = 1 if 'pode_ver_saldo' in request.form else 0

            # Verifica se o usuário já existe
            cursor.execute("SELECT id FROM usuarios WHERE username = ? AND D_E_L_E_T_ <> '*'", username)
            if cursor.fetchone():
                flash('Nome de usuário já existe. Por favor, escolha outro.', 'error')
            else:
                hashed_password = generate_password_hash(password)
                cursor.execute(
                    "INSERT INTO usuarios (username, password_hash, filial, role_id, pode_ver_saldo) VALUES (?, ?, ?, ?, ?)",
                    username, hashed_password, filial, role_id, pode_ver_saldo
                )
                conn.commit()
                flash('Usuário criado com sucesso!', 'success')
                return redirect(url_for('admin'))

        # Lógica para exibir a página
        # 1. Obter a lista de perfis (roles) para o dropdown
        cursor.execute("SELECT id, name FROM roles WHERE D_E_L_E_T_ = '' ORDER BY name")
        roles = cursor.fetchall()
        
        # 2. Obter a lista de usuários para a tabela, incluindo o nome do perfil
        cursor.execute("""
            SELECT u.id, u.username, u.filial, u.pode_ver_saldo, r.name AS role_name
            FROM usuarios u
            LEFT JOIN roles r ON u.role_id = r.id
            WHERE u.D_E_L_E_T_ <> '*'
            ORDER BY u.username
        """)
        users = cursor.fetchall()
        
        # Obtém a lista de filiais para o filtro (se você tiver)
        cursor.execute("SELECT DISTINCT filial FROM usuarios WHERE D_E_L_E_T_ <> '*' ORDER BY filial")
        filiais = cursor.fetchall()

    except pyodbc.Error as ex:
        sqlstate = ex.args[0]
        flash(f"Erro ao acessar o banco de dados: {sqlstate}", 'error')
        roles = []
        users = []
        filiais = []
    finally:
        if conn:
            conn.close()

    return render_template('admin.html', roles=roles, users=users, filiais=filiais, username=session.get('username'))

@app.route('/admin/edit_user/<int:user_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_user(user_id):
    conn = None
    try:
        conn = get_db_connection(DB_INVENTARIO)
        cursor = conn.cursor()

        # Busca os dados do usuário a ser editado
        cursor.execute("SELECT id, username, filial, role_id, pode_ver_saldo FROM usuarios WHERE id = ? AND D_E_L_E_T_ <> '*'", user_id)
        user = cursor.fetchone()
        
        if not user:
            flash("Usuário não encontrado.", 'error')
            return redirect(url_for('admin'))
            
        # Busca a lista de perfis para o dropdown
        cursor.execute("SELECT id, name FROM roles WHERE D_E_L_E_T_ = '' ORDER BY name")
        roles = cursor.fetchall()
        
        if request.method == 'POST':
            # Processa a submissão do formulário
            new_username = request.form['username']
            new_filial = request.form['filial']
            new_role_id = request.form['role_id']
            pode_ver_saldo = 1 if 'pode_ver_saldo' in request.form else 0
            
            # Atualiza o usuário no banco de dados
            cursor.execute(
                "UPDATE usuarios SET username = ?, filial = ?, role_id = ?, pode_ver_saldo = ? WHERE id = ?",
                new_username, new_filial, new_role_id, pode_ver_saldo, user_id
            )
            conn.commit()
            flash('Usuário atualizado com sucesso!', 'success')
            return redirect(url_for('admin'))
            
        # Se o método for GET, renderiza o template de edição
        return render_template(
            'edit_user.html', 
            user=user, 
            roles=roles, 
            username=session.get('username')
        )
    except pyodbc.Error as ex:
        sqlstate = ex.args[0]
        flash(f"Erro ao acessar o banco de dados: {sqlstate}", 'error')
        return redirect(url_for('admin'))
    finally:
        if conn:
            conn.close()

# --- NOVAS ROTAS PARA GESTÃO DE PERFIS ---

@app.route('/admin/roles')
@login_required
def manage_roles():
    conn = get_db_connection(DB_INVENTARIO)
    cursor = conn.cursor()
    cursor.execute("SELECT id, name FROM roles WHERE D_E_L_E_T_ <> '*' ORDER BY name")
    roles = cursor.fetchall()
    conn.close()
    return render_template('roles.html', roles=roles, username=session.get('username'))

@app.route('/admin/roles/edit/<int:role_id>', methods=['GET', 'POST'])
@login_required
def edit_role(role_id):
    conn = get_db_connection(DB_INVENTARIO)
    cursor = conn.cursor()

    # --- CORREÇÃO PERMANENTE ---
    # Impede a edição do perfil de Administrador (ID 1)
    if role_id == 1:
        flash("O perfil de Administrador não pode ser editado. Ele possui todas as permissões por padrão.", 'info')
        conn.close()
        return redirect(url_for('manage_roles'))
    
    if request.method == 'POST':
        new_name = request.form.get('name')
        permissions = request.form.getlist('permissions')
        
        # Atualiza o nome do perfil
        cursor.execute("UPDATE roles SET name = ? WHERE id = ?", new_name, role_id)
        
        # Atualiza as permissões
        cursor.execute("DELETE FROM role_permissions WHERE role_id = ?", role_id)
        if permissions:
            permission_data = [(role_id, p) for p in permissions]
            cursor.executemany("INSERT INTO role_permissions (role_id, menu_endpoint) VALUES (?, ?)", permission_data)
        
        conn.commit()
        flash(f"Perfil '{new_name}' atualizado com sucesso!", 'success')
        conn.close()
        return redirect(url_for('manage_roles'))

    cursor.execute("SELECT id, name FROM roles WHERE id = ? AND D_E_L_E_T_ <> '*'", role_id)
    role = cursor.fetchone()

    # --- AQUI ESTÁ A CORREÇÃO ---
    # Se o perfil for o Admin (ID 1), exibe todas as permissões
    if role and role.id == 1:
        current_permissions = set(AVAILABLE_MENUS.keys())
    else:
        cursor.execute("SELECT menu_endpoint FROM role_permissions WHERE role_id = ?", role_id)
        current_permissions = {row.menu_endpoint for row in cursor.fetchall()}
    
    conn.close()
    if not role:
        flash("Perfil não encontrado.", 'error')
        return redirect(url_for('manage_roles'))
        
    return render_template('edit_role.html', role=role, 
                           available_menus=AVAILABLE_MENUS, 
                           current_permissions=current_permissions, 
                           username=session.get('username'))


@app.route('/admin/reset_password/<int:user_id>', methods=['POST'])
@login_required
@role_required('Admin')
def reset_password(user_id):
    new_password = request.form.get('password')
    if not new_password:
        flash('A nova senha não pode estar em branco.', 'error')
        return redirect(url_for('edit_user', user_id=user_id))
    
    hashed_password = generate_password_hash(new_password)
    conn = get_db_connection(DB_INVENTARIO)
    cursor = conn.cursor()
    cursor.execute("UPDATE usuarios SET password_hash = ? WHERE id = ?", hashed_password, user_id)
    conn.commit()
    conn.close()
    
    flash('Senha do usuário resetada com sucesso!', 'success')
    return redirect(url_for('edit_user', user_id=user_id))

@app.route('/admin/delete_user/<int:user_id>', methods=['POST'])
@login_required
@role_required('Admin')
def delete_user(user_id):
    conn = get_db_connection(DB_INVENTARIO)
    cursor = conn.cursor()
    
    cursor.execute("SELECT username FROM usuarios WHERE id = ?", user_id)
    user_to_delete = cursor.fetchone()
    if user_to_delete.username == 'admin':
        flash('Não é permitido excluir o usuário "admin" principal.', 'error')
        return redirect(url_for('admin'))
    if user_id == session.get('user_id'):
        flash('Você não pode excluir a si mesmo.', 'error')
        return redirect(url_for('admin'))

    cursor.execute("UPDATE usuarios SET D_E_L_E_T_ = '*' WHERE id = ?", user_id)
    conn.commit()
    conn.close()
    flash('Usuário excluído com sucesso!', 'success')
    return redirect(url_for('admin'))

@app.route('/inventarios', methods=['GET', 'POST'])
@login_required
@role_required('Admin', 'Supervisor', 'Operador', 'Inventarista')
def gerenciar_inventarios():
    conn = get_db_connection(DB_INVENTARIO)
    cursor = conn.cursor()

    if request.method == 'POST':
        descricao = request.form['descricao']
        if descricao:
            cursor.execute(
                "INSERT INTO inventario_sessoes (descricao, usuario_id_abertura) VALUES (?, ?)",
                descricao, session['user_id']
            )
            conn.commit()
        return redirect(url_for('gerenciar_inventarios'))

    query = """
    SELECT s.id, s.descricao, s.data_abertura, s.status, u.username 
    FROM inventario_sessoes s
    JOIN usuarios u ON s.usuario_id_abertura = u.id
    WHERE s.D_E_L_E_T_ <> '*'
    ORDER BY s.data_abertura DESC
    """
    cursor.execute(query)
    inventarios = cursor.fetchall()
    conn.close()
    return render_template('inventarios.html', inventarios=inventarios, username=session.get('username'))

@app.route('/inventarios/fechar/<int:sessao_id>', methods=['POST'])
@login_required
@role_required('Admin', 'Supervisor')
def fechar_inventario(sessao_id):
    conn = get_db_connection(DB_INVENTARIO)
    cursor = conn.cursor()

    try:
        # --- ETAPA 1: Limpeza de Status 'Recontagem Solicitada' ---
        # Identifica todos os itens que já tiveram uma contagem mais recente
        cursor.execute("""
            WITH LatestCounts AS (
                SELECT codigo_item, MAX(numero_contagem) as max_num
                FROM contagens
                WHERE sessao_id = ?
                GROUP BY codigo_item
            )
            UPDATE c
            SET c.status_contagem = 'Finalizado'
            FROM contagens c
            JOIN LatestCounts lc ON c.codigo_item = lc.codigo_item
            WHERE c.sessao_id = ?
              AND c.numero_contagem < lc.max_num
              AND c.status_contagem = 'Recontagem Solicitada'
        """, sessao_id, sessao_id)
        conn.commit() # Efetiva a limpeza antes da verificação

        # --- ETAPA 2: Verificação de Pendências Reais ---
        # Agora, a verificação só vai acusar itens que estão genuinamente pendentes
        cursor.execute("SELECT COUNT(*) FROM contagens WHERE sessao_id = ? AND status_contagem IN ('Pendente', 'Recontagem Solicitada')", sessao_id)
        pendencias = cursor.fetchone()[0]

        if pendencias > 0:
            flash(f"Não foi possível fechar o inventário. Existem {pendencias} item(s) com contagem genuinamente pendente.", 'error')
            conn.close()
            return redirect(url_for('gerenciar_inventarios'))
        
        # --- ETAPA 3: Validação final e fechamento ---
        # A lógica de validação de 3 contagens e ignorados (se aplicável)
        dados_analiticos = _gerar_dados_analiticos(sessao_id)
        if dados_analiticos and dados_analiticos['contagens_agrupadas']:
             for grupo_contagens in dados_analiticos['contagens_agrupadas']:
                for contagem in grupo_contagens:
                    if contagem['status_contagem'] in ['Validado', 'Ignorado']:
                        cursor.execute("UPDATE contagens SET status_contagem = ? WHERE id = ?", contagem['status_contagem'], contagem['id'])
        
        # Finalmente, fecha o inventário
        cursor.execute(
            "UPDATE inventario_sessoes SET status = 'Fechado', data_fechamento = GETDATE() WHERE id = ? AND status = 'Aberto'",
            sessao_id
        )
        conn.commit()
        flash('Inventário fechado com sucesso!', 'success')

    except Exception as e:
        conn.rollback()
        flash(f"Ocorreu um erro inesperado ao fechar o inventário: {e}", 'error')
    finally:
        if conn:
            conn.close()
            
    return redirect(url_for('gerenciar_inventarios'))

@app.route('/inventarios/edit/<int:sessao_id>', methods=['GET', 'POST'])
@login_required
@role_required('Admin', 'Supervisor')
def edit_inventario(sessao_id):
    conn = get_db_connection(DB_INVENTARIO)
    cursor = conn.cursor()

    if request.method == 'POST':
        nova_descricao = request.form.get('descricao')
        if nova_descricao:
            cursor.execute("UPDATE inventario_sessoes SET descricao = ? WHERE id = ?", nova_descricao, sessao_id)
            conn.commit()
            flash('Descrição do inventário atualizada com sucesso!', 'success')
        else:
            flash('A descrição não pode estar em branco.', 'error')
        conn.close()
        return redirect(url_for('gerenciar_inventarios'))

    cursor.execute("SELECT id, descricao FROM inventario_sessoes WHERE id = ? AND D_E_L_E_T_ <> '*'", sessao_id)
    inventario = cursor.fetchone()
    conn.close()
    if not inventario:
        return redirect(url_for('gerenciar_inventarios'))
    return render_template('edit_inventario.html', inventario=inventario, username=session.get('username'))

@app.route('/inventarios/reabrir/<int:sessao_id>', methods=['POST'])
@login_required
@role_required('Admin', 'Supervisor')
def reabrir_inventario(sessao_id):
    conn = get_db_connection(DB_INVENTARIO)
    cursor = conn.cursor()
    cursor.execute("UPDATE inventario_sessoes SET status = 'Aberto', data_fechamento = NULL WHERE id = ?", sessao_id)
    conn.commit()
    conn.close()
    flash('Inventário reaberto com sucesso!', 'success')
    return redirect(url_for('gerenciar_inventarios'))

@app.route('/inventarios/excluir/<int:sessao_id>', methods=['POST'])
@login_required
@role_required('Admin')
def excluir_inventario(sessao_id):
    conn = get_db_connection(DB_INVENTARIO)
    cursor = conn.cursor()
    try:
        cursor.execute("UPDATE inventario_sessoes SET D_E_L_E_T_ = '*' WHERE id = ?", sessao_id)
        conn.commit()
        flash('Inventário excluído com sucesso!', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Ocorreu um erro ao excluir o inventário: {e}', 'error')
    finally:
        conn.close()
    return redirect(url_for('gerenciar_inventarios'))

# ======================================================================================
# ========= FUNÇÃO CORRIGIDA - INÍCIO =================================================
# ======================================================================================
@app.route('/inventario/<int:sessao_id>')
@login_required
@role_required('Admin', 'Supervisor', 'Operador', 'Inventarista')
def visualizar_inventario(sessao_id):
    """
    CORREÇÃO: Esta função agora simplesmente chama a função de análise de dados
    e passa os resultados diretamente para o template, sem lógicas duplicadas.
    """
    dados = _gerar_dados_analiticos(sessao_id)

    if not dados:
        flash("Inventário não encontrado ou não possui contagens.", "error")
        return redirect(url_for('gerenciar_inventarios'))

    return render_template('visualizar_inventario.html',
                           inventario=dados['inventario'],
                           contagens_agrupadas=dados['contagens_agrupadas'],
                           username=session.get('username'))
# ======================================================================================
# ========= FUNÇÃO CORRIGIDA - FIM ===================================================
# ======================================================================================

@app.route('/inventarios/exportar_csv/<int:sessao_id>')
@login_required
@role_required('Admin', 'Supervisor')
def exportar_csv(sessao_id):
    conn = get_db_connection(DB_INVENTARIO)
    cursor = conn.cursor()
    
    query = f"""
    SELECT 
        s.id as sessao_id, s.descricao, c.data_contagem, c.codigo_item, 
        ISNULL(b1.B1_DESC, 'N/A') as B1_DESC,
        c.filial, c.local,
        c.saldo_sistema, c.quantidade_contada, (c.quantidade_contada - c.saldo_sistema) as diferenca,
        u.username as usuario_contagem
    FROM contagens c
    JOIN inventario_sessoes s ON c.sessao_id = s.id
    JOIN usuarios u ON c.usuario_id = u.id
    LEFT JOIN {DB_BI}.dbo.SB1010 b1 ON c.codigo_item COLLATE DATABASE_DEFAULT = b1.B1_COD COLLATE DATABASE_DEFAULT AND b1.D_E_L_E_T_ <> '*'
    WHERE c.sessao_id = ?
    ORDER BY c.codigo_item, c.filial, c.local
    """
    cursor.execute(query, sessao_id)
    contagens = cursor.fetchall()

    if not contagens:
        conn.close()
        flash('Nenhuma contagem para exportar.', 'info')
        return redirect(url_for('gerenciar_inventarios'))

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Inventario_ID', 'Descricao_Inventario', 'Data_Contagem', 'Codigo_Item', 'Descricao_Item', 'Filial', 'Local', 'Saldo_Sistema', 'Qtd_Contada', 'Diferenca', 'Usuario_Contagem'])
    
    for row in contagens:
        # ===== ALTERAÇÃO DA FORMATAÇÃO AQUI =====
        writer.writerow([
            row.sessao_id, row.descricao, row.data_contagem.strftime('%Y-%m-%d %H:%M:%S'),
            row.codigo_item, row.B1_DESC.strip(), row.filial, row.local,
            f"{row.saldo_sistema:.2f}",
            f"{row.quantidade_contada:.2f}",
            f"{row.diferenca:.2f}",
            row.usuario_contagem
        ])
    
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = f"attachment; filename=inventario_{sessao_id}.csv"
    response.headers["Content-type"] = "text/csv; charset=utf-8" # Adicionado charset para melhor compatibilidade
    conn.close()
    return response


@app.route('/inventarios/exportar_excel/<int:sessao_id>')
@login_required
@role_required('Admin', 'Supervisor')
def exportar_excel(sessao_id):
    dados = _gerar_dados_analiticos(sessao_id)
    if not dados:
        flash("Nenhuma contagem para exportar.", 'info')
        return redirect(url_for('selecionar_inventario_analitico'))

    wb = Workbook()
    
    # --- Aba de Resumo ---
    ws_resumo = wb.active
    ws_resumo.title = "Resumo Gerencial"
    
    ws_resumo.append(['Indicador', 'Valor'])
    ws_resumo.append(['Acurácia do Inventário', f"{dados['metricas']['Acuracia']:.2f}%"])
    ws_resumo.append(['Valor do Estoque Apurado', dados['metricas']['ValorApurado']])
    ws_resumo.append(['Valor Saldo Sistema', dados['metricas']['ValorTotalSistema']])
    ws_resumo.append(['Valor Total em Sobra', dados['metricas']['ValorSobra']])
    ws_resumo.append(['Valor Total em Falta', dados['metricas']['ValorFalta']])
    
    ws_resumo.append([' '])
    ws_resumo.append(['Divergência por Grupo', 'Valor em Sobra', 'Valor em Falta'])
    for grupo, valores in dados['divergencia_por_grupo'].items():
        ws_resumo.append([grupo, valores['sobra'], valores['falta']])

    ws_resumo.append([' '])
    ws_resumo.append(['Top 5 Maiores Faltas (em Valor)', 'Código', 'Descrição'])
    for item in dados['top_faltas']:
        ws_resumo.append([item['valor'], item['codigo'], item['descricao']])

    ws_resumo.append([' '])
    ws_resumo.append(['Top 5 Maiores Sobras (em Valor)', 'Código', 'Descrição'])
    for item in dados['top_sobras']:
        ws_resumo.append([item['valor'], item['codigo'], item['descricao']])

    # --- Aba de Detalhes ---
    ws_detalhes = wb.create_sheet(title="Detalhes das Contagens")
    headers = ['Código', 'Descrição', 'Filial', 'Local', 'Custo Unit.', 'Saldo Sistema (Qtd)', 'Qtd. Contada', 'Diferença (Qtd)', 'Valor Diferença', 'Usuário', 'Num. Contagem']
    ws_detalhes.append(headers)
    
    for item_processed in dados['contagens_agrupadas']:
        for location_group in item_processed['locations']:
            for contagem in location_group:
                valor_diferenca = contagem['diferenca'] * contagem['custo_unitario']
                ws_detalhes.append([
                    contagem['codigo_item'], contagem['B1_DESC'], contagem['filial'], contagem['local'],
                    contagem['custo_unitario'], contagem['saldo_sistema'], contagem['quantidade_contada'],
                    contagem['diferenca'], valor_diferenca, contagem['username'], contagem['numero_contagem']
                ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment;filename=analise_inventario_{sessao_id}.xlsx'})




@app.route('/contagem_planejada', methods=['GET', 'POST'])
@login_required
def contagem_planejada():
    itens = None
    data_inicial_form = request.form.get('data_inicial', '')
    data_final_form = request.form.get('data_final', '')
    grupos_selecionados_form = request.form.getlist('grupos')
    sessao_selecionada_form = request.form.get('sessao_id', '')
    filiais_selecionadas_form = request.form.getlist('filiais')
    locais_selecionados_form = request.form.getlist('locais')
    
    conn_inventario = get_db_connection(DB_INVENTARIO)
    cursor_inventario = conn_inventario.cursor()
    cursor_inventario.execute("SELECT id, descricao FROM inventario_sessoes WHERE status = 'Aberto' AND D_E_L_E_T_ <> '*' ORDER BY data_abertura DESC")
    sessoes_abertas = cursor_inventario.fetchall()
    conn_inventario.close()

    conn_bi = get_db_connection(DB_BI)
    cursor_bi = conn_bi.cursor()
    cursor_bi.execute("SELECT DISTINCT B2_FILIAL FROM SB2010 WHERE D_E_L_E_T_ <> '*' ORDER BY B2_FILIAL")
    filiais = [row.B2_FILIAL.strip() for row in cursor_bi.fetchall()]
    cursor_bi.execute("SELECT DISTINCT B2_LOCAL FROM SB2010 WHERE D_E_L_E_T_ <> '*' ORDER BY B2_LOCAL")
    locais = [row.B2_LOCAL.strip() for row in cursor_bi.fetchall()]
    conn_bi.close()

    if request.method == 'POST':
        if data_inicial_form and data_final_form and sessao_selecionada_form:
            data_inicial_db = data_inicial_form.replace('-', '')
            data_final_db = data_final_form.replace('-', '')

            conn_bi_post = get_db_connection(DB_BI)
            cursor_bi_post = conn_bi_post.cursor()
            
            query_parts = [
                f"SELECT b2.B2_COD, MAX(b1.B1_DESC) as B1_DESC, MAX(b2.B2_DMOV) as UltimaMov,",
                f"CASE WHEN COUNT(c.id) > 0 THEN 'Contado' ELSE 'Pendente' END as StatusContagem",
                f"FROM SB2010 b2",
                f"LEFT JOIN SB1010 b1 ON b2.B2_COD = b1.B1_COD AND b1.D_E_L_E_T_ <> '*'",
                f"LEFT JOIN {DB_INVENTARIO}.dbo.contagens c ON b2.B2_COD COLLATE DATABASE_DEFAULT = c.codigo_item COLLATE DATABASE_DEFAULT AND c.sessao_id = ?",
                f"WHERE b2.B2_DMOV BETWEEN ? AND ? AND b2.D_E_L_E_T_ <> '*'"
            ]
            params = [sessao_selecionada_form, data_inicial_db, data_final_db]

            if grupos_selecionados_form:
                placeholders = ','.join(['?'] * len(grupos_selecionados_form))
                query_parts.append(f"AND SUBSTRING(b2.B2_COD, 1, 4) IN ({placeholders})")
                params.extend(grupos_selecionados_form)
            
            if filiais_selecionadas_form:
                placeholders = ','.join(['?'] * len(filiais_selecionadas_form))
                query_parts.append(f"AND b2.B2_FILIAL IN ({placeholders})")
                params.extend(filiais_selecionadas_form)

            if locais_selecionados_form:
                placeholders = ','.join(['?'] * len(locais_selecionados_form))
                query_parts.append(f"AND b2.B2_LOCAL IN ({placeholders})")
                params.extend(locais_selecionados_form)

            query_parts.extend([
                "GROUP BY b2.B2_COD",
                "ORDER BY UltimaMov ASC"
            ])
            
            query = ' '.join(query_parts)
            
            cursor_bi_post.execute(query, *params)
            itens = cursor_bi_post.fetchall()
            conn_bi_post.close()

    return render_template('contagem_planejada.html', 
                           itens=itens, 
                           sessoes=sessoes_abertas,
                           filiais=filiais,
                           locais=locais,
                           sessao_selecionada=sessao_selecionada_form,
                           grupos_selecionados=grupos_selecionados_form,
                           filiais_selecionadas=filiais_selecionadas_form,
                           locais_selecionados=locais_selecionados_form,
                           data_inicial_pesquisada=data_inicial_form,
                           data_final_pesquisada=data_final_form,
                           username=session.get('username'))

# ======================================================================================
# ========= FUNÇÃO DE ANÁLISE REFEITA - INÍCIO ========================================
# ======================================================================================
def _gerar_dados_analiticos(sessao_id):
    conn = get_db_connection(DB_INVENTARIO)
    cursor = conn.cursor()
    inventario = cursor.execute("SELECT id, descricao, data_abertura, data_fechamento, status FROM inventario_sessoes WHERE id = ?", sessao_id).fetchone()
    if not inventario:
        conn.close()
        return None

    query_contagens = "SELECT * FROM contagens WHERE sessao_id = ?"
    contagens_da_sessao = cursor.execute(query_contagens, sessao_id).fetchall()
    
    if not contagens_da_sessao:
        conn.close()
        return {
            "inventario": inventario, "contagens_agrupadas": [],
            "metricas": {}, "chart_labels": [], "chart_data": [], "top_faltas": [], "top_sobras": [], "divergencia_por_grupo": {}
        }

    user_ids = list(set(c.usuario_id for c in contagens_da_sessao))
    users = {}
    if user_ids:
        placeholders = ','.join(['?'] * len(user_ids))
        cursor.execute(f"SELECT id, username FROM usuarios WHERE id IN ({placeholders})", *user_ids)
        for row in cursor.fetchall():
            users[row.id] = row.username
    conn.close()

    contagens_agrupadas_por_item = {}
    for c in contagens_da_sessao:
        c_dict = dict(zip([column[0] for column in c.cursor_description], c))
        c_dict['username'] = users.get(c_dict['usuario_id'], 'N/A')
        c_dict['diferenca'] = c_dict['quantidade_contada'] - c_dict['saldo_sistema']
        codigo_item = c_dict['codigo_item']
        if codigo_item not in contagens_agrupadas_por_item:
            contagens_agrupadas_por_item[codigo_item] = []
        contagens_agrupadas_por_item[codigo_item].append(c_dict)

    codigos_itens = list(contagens_agrupadas_por_item.keys())
    placeholders = ','.join(['?'] * len(codigos_itens))
    query_dados_itens = f"SELECT B1_COD, B1_DESC FROM {DB_BI}.dbo.SB1010 WHERE B1_COD IN ({placeholders}) AND D_E_L_E_T_ <> '*'"
    conn_bi = get_db_connection(DB_BI)
    cursor_bi = conn_bi.cursor()
    cursor_bi.execute(query_dados_itens, *codigos_itens)
    dados_itens_raw = cursor_bi.fetchall()
    conn_bi.close()
    descricoes = {row.B1_COD.strip(): row.B1_DESC.strip() for row in dados_itens_raw}

    contagens_processadas = []
    for item_code, contagens_do_item in contagens_agrupadas_por_item.items():
        for c in contagens_do_item:
            c['B1_DESC'] = descricoes.get(item_code, 'N/A')

        locais_do_item = {}
        for c in contagens_do_item:
            chave_local = (c['filial'], c['local'])
            if chave_local not in locais_do_item:
                locais_do_item[chave_local] = []
            locais_do_item[chave_local].append(c)

        max_n_contagem_geral = max(c['numero_contagem'] for c in contagens_do_item)
        
        ainda_pendente_geral = any(
            c['status_contagem'] == 'Pendente' 
            for c in contagens_do_item 
            if c['numero_contagem'] == max_n_contagem_geral
        )
        
        divergencia_geral_encontrada = False
        if not ainda_pendente_geral:
            divergencia_geral_encontrada = any(
                c['diferenca'] != 0 
                for c in contagens_do_item 
                if c['numero_contagem'] == max_n_contagem_geral
            )

        necessita_recontagem = divergencia_geral_encontrada and max_n_contagem_geral < 3
        limite_atingido = divergencia_geral_encontrada and max_n_contagem_geral >= 3
        tudo_ok = not divergencia_geral_encontrada and not ainda_pendente_geral

        for loc_group in locais_do_item.values():
            loc_group.sort(key=lambda x: x['numero_contagem'])

        contagens_processadas.append({
            'item_info': contagens_do_item[0],
            'locations': list(locais_do_item.values()),
            'necessita_recontagem': necessita_recontagem,
            'proximo_numero_contagem': max_n_contagem_geral + 1,
            'id_para_recontagem': contagens_do_item[-1]['id'],
            'limite_atingido': limite_atingido,
            'tudo_ok': tudo_ok,
            'pendente_geral': ainda_pendente_geral
        })

    return {
        "inventario": inventario,
        "contagens_agrupadas": contagens_processadas
    }
# ======================================================================================
# ========= FUNÇÃO DE ANÁLISE REFEITA - FIM ===========================================
# ======================================================================================


@app.route('/dashboard_analitico')
@login_required
@role_required('Admin', 'Supervisor')
def selecionar_inventario_analitico():
    conn = get_db_connection(DB_INVENTARIO)
    cursor = conn.cursor()
    cursor.execute("SELECT id, descricao, data_fechamento FROM inventario_sessoes WHERE status = 'Fechado' AND D_E_L_E_T_ <> '*' ORDER BY data_fechamento DESC")
    inventarios_fechados = cursor.fetchall()
    conn.close()
    return render_template('selecionar_analise.html', inventarios=inventarios_fechados, username=session.get('username'))

@app.route('/dashboard_analitico/<int:sessao_id>')
@login_required
@role_required('Admin', 'Supervisor')
def dashboard_analitico(sessao_id):
    dados = _gerar_dados_analiticos(sessao_id)
    print("\n" + "="*50)
    print(f"[DEBUG] Iniciando _gerar_dados_analiticos para Sessão ID: {sessao_id}")
    print("="*50)
    if not dados:
        flash("Inventário não encontrado.", "error")
        return redirect(url_for('selecionar_inventario_analitico'))
    
    return render_template('dashboard_analitico.html', **dados, username=session.get('username'))



@app.route('/inventario/gerar_pdf/<int:sessao_id>')
@login_required
@role_required('Admin', 'Supervisor')
def gerar_relatorio_pdf(sessao_id):
    dados = _gerar_dados_analiticos(sessao_id)
    if not dados:
        flash("Inventário não encontrado.", "error")
        return redirect(url_for('selecionar_inventario_analitico'))
            
    # Lógica do Gráfico (sem alteração)
    plt.figure(figsize=(5, 5))
    metricas = dados['metricas']
    total_corretos = (metricas.get('TotalContagens', 0)) - (metricas.get('TotalDiferencas', 0))
    data_pie = [total_corretos, metricas.get('TotalSobra', 0), metricas.get('TotalFalta', 0)]
    labels_pie = ['Corretos', 'Com Sobra', 'Com Falta']
    colors_pie = ['#05B540', '#FFAA00', '#FF6600']
    plt.pie(data_pie, labels=labels_pie, autopct='%1.1f%%', colors=colors_pie, startangle=90)
    plt.title('Análise de Diferenças (em Qtd)')
    img_pie = BytesIO()
    plt.savefig(img_pie, format='png', bbox_inches='tight')
    img_pie.seek(0)
    pie_chart_url = "data:image/png;base64," + base64.b64encode(img_pie.getvalue()).decode('utf8')
    plt.close()

    # --- NOVA LÓGICA PARA CARREGAR O LOGO LOCAL ---
    logo_url = ''
    try:
        path = os.path.join(app.root_path, 'static', 'logo', 'logo1.png')
        with open(path, 'rb') as logo_file:
            encoded_logo = base64.b64encode(logo_file.read()).decode('utf-8')
            logo_url = f"data:image/png;base64,{encoded_logo}"
    except Exception as e:
        print(f"Erro ao carregar o logo: {e}") # Opcional: logar o erro
    # --- FIM DA NOVA LÓGICA ---

    html_renderizado = render_template('report_template.html', **dados, pie_chart_url=pie_chart_url, logo_url=logo_url)
    
    pdf = HTML(string=html_renderizado).write_pdf()
    return Response(pdf, mimetype='application/pdf', headers={'Content-Disposition': f'attachment;filename=relatorio_inventario_{sessao_id}.pdf'})


@app.route('/inventario/enviar_email/<int:sessao_id>', methods=['POST'])
@login_required
@role_required('Admin', 'Supervisor')
def enviar_relatorio_email(sessao_id):
    recipient_email = request.form.get('recipient_email')
    if not recipient_email:
        flash('O e-mail do destinatário é obrigatório.', 'error')
        return redirect(url_for('dashboard_analitico', sessao_id=sessao_id))

    dados = _gerar_dados_analiticos(sessao_id)
    if not dados:
        flash("Inventário não encontrado para gerar o e-mail.", "error")
        return redirect(url_for('dashboard_analitico', sessao_id=sessao_id))

    # Lógica do Gráfico (idêntica à da geração de PDF)
    plt.figure(figsize=(5, 5))
    metricas = dados['metricas']
    total_corretos = (metricas.get('TotalContagens', 0)) - (metricas.get('TotalDiferencas', 0))
    data_pie = [total_corretos, metricas.get('TotalSobra', 0), metricas.get('TotalFalta', 0)]
    labels_pie = ['Corretos', 'Com Sobra', 'Com Falta']
    colors_pie = ['#05B540', '#FFAA00', '#FF6600']
    plt.pie(data_pie, labels=labels_pie, autopct='%1.1f%%', colors=colors_pie, startangle=90)
    plt.title('Análise de Diferenças (em Qtd)')
    img_pie = BytesIO()
    plt.savefig(img_pie, format='png', bbox_inches='tight')
    img_pie.seek(0)
    pie_chart_url = "data:image/png;base64," + base64.b64encode(img_pie.getvalue()).decode('utf8')
    plt.close()

    # --- LÓGICA PARA CARREGAR O LOGO LOCAL (REPETIDA AQUI) ---
    logo_url = ''
    try:
        path = os.path.join(app.root_path, 'static', 'logo', 'logo1.png')
        with open(path, 'rb') as logo_file:
            encoded_logo = base64.b64encode(logo_file.read()).decode('utf-8')
            logo_url = f"data:image/png;base64,{encoded_logo}"
    except Exception as e:
        print(f"Erro ao carregar o logo: {e}")
    # --- FIM DA LÓGICA DO LOGO ---

    html_renderizado = render_template('report_template.html', **dados, pie_chart_url=pie_chart_url, logo_url=logo_url)
    pdf_content = HTML(string=html_renderizado).write_pdf()
    
    message = MIMEMultipart()
    message["From"] = EMAIL_SENDER
    message["To"] = recipient_email
    message["Subject"] = f"Relatório do Inventário: {dados['inventario'].descricao}"
    
    body_html = f"<html><body><p>Olá,</p><p>Segue em anexo o relatório em PDF para o inventário '<b>{dados['inventario'].descricao}</b>'.</p><br><p>Atenciosamente,</p><p>Sistema de Inventário</p></body></html>"
    message.attach(MIMEText(body_html, "html"))

    part = MIMEApplication(pdf_content, Name=f"relatorio_inventario_{sessao_id}.pdf")
    part['Content-Disposition'] = f'attachment; filename="relatorio_inventario_{sessao_id}.pdf"'
    message.attach(part)

    try:
        context = ssl.create_default_context()
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls(context=context)
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)
            server.sendmail(EMAIL_SENDER, recipient_email, message.as_string())
        flash(f'Relatório enviado com sucesso para {recipient_email}!', 'success')
    except Exception as e:
        flash(f'Ocorreu um erro ao enviar o e-mail: {e}', 'error')
        print(f"Erro de e-mail: {e}")

    return redirect(url_for('dashboard_analitico', sessao_id=sessao_id))

@app.route('/recontagem')
@login_required
def lista_recontagem():
    conn = get_db_connection(DB_INVENTARIO)
    cursor = conn.cursor()
    
    query = f"""
    SELECT 
        c.codigo_item, c.filial, s.descricao as inventario_desc, 
        s.id as sessao_id, ISNULL(b1.B1_DESC, 'N/A') as B1_DESC
    FROM contagens c
    JOIN inventario_sessoes s ON c.sessao_id = s.id
    LEFT JOIN {DB_BI}.dbo.SB1010 b1 ON c.codigo_item COLLATE DATABASE_DEFAULT = b1.B1_COD COLLATE DATABASE_DEFAULT AND b1.D_E_L_E_T_ <> '*'
    WHERE c.status_contagem = 'Pendente' AND s.status = 'Aberto' AND s.D_E_L_E_T_ <> '*'
    ORDER BY s.id, c.codigo_item
    """
    cursor.execute(query)
    itens_recontar_raw = cursor.fetchall()

    # Agrupa as tarefas pendentes para que cada item apareça apenas uma vez
    itens_agrupados = {}
    for item in itens_recontar_raw:
        chave = (item.sessao_id, item.codigo_item.strip())
        if chave not in itens_agrupados:
            itens_agrupados[chave] = {
                'codigo_item': item.codigo_item.strip(),
                'B1_DESC': item.B1_DESC.strip(),
                'inventario_desc': item.inventario_desc,
                'sessao_id': item.sessao_id,
                'filial': item.filial, # Guarda a primeira filial para encontrar o grupo
                'local': '', # Não é necessário para a exibição agrupada
                'is_consolidated': False,
                'group_name': None
            }
    
    # Itera sobre os itens únicos para definir a flag de consolidação
    itens_recontar_final = []
    for chave, item_data in itens_agrupados.items():
        sessao_id, codigo_item = chave
        
        # Verifica quantas linhas a contagem original (nº 1) tinha. Se for mais de 1, é consolidada.
        cursor.execute("SELECT COUNT(id) FROM contagens WHERE sessao_id = ? AND codigo_item = ? AND numero_contagem = 1", sessao_id, codigo_item)
        original_count_rows = cursor.fetchone()[0]
        
        if original_count_rows > 1:
            item_data['is_consolidated'] = True

        # Determina o nome do grupo para o link
        if item_data.get('filial'):
            for g_name, filiais in ALMOXARIFADOS_GRUPO.items():
                if item_data['filial'] in filiais:
                    item_data['group_name'] = g_name
                    break
        
        itens_recontar_final.append(item_data)
    
    conn.close()
    return render_template('recontagem.html', itens=itens_recontar_final, username=session.get('username'))

@app.route('/inventario/solicitar_recontagem/<int:contagem_id>', methods=['POST'])
@login_required
@role_required('Admin', 'Supervisor')
def solicitar_recontagem(contagem_id):
    conn = get_db_connection(DB_INVENTARIO)
    cursor = conn.cursor()
    sessao_id_redirect = None
    try:
        # Pega a contagem base (específica)
        cursor.execute("SELECT id, sessao_id, codigo_item, filial, local, numero_contagem FROM contagens WHERE id = ?", contagem_id)
        contagem_base = cursor.fetchone()

        if not contagem_base:
            flash('Contagem original não encontrada.', 'error')
            return redirect(request.referrer or url_for('gerenciar_inventarios'))

        sessao_id = contagem_base.sessao_id
        sessao_id_redirect = sessao_id
        codigo_item = contagem_base.codigo_item
        filial = contagem_base.filial
        local = contagem_base.local
        ultimo_num_contagem = contagem_base.numero_contagem
        novo_num_contagem = ultimo_num_contagem + 1

        if novo_num_contagem > 3:
            flash('Limite de 3 contagens atingido para este item neste local.', 'warning')
            return redirect(url_for('visualizar_inventario', sessao_id=sessao_id))

        # Recupera a contagem "pai" (primeira contagem) para esse sessao_id + codigo_item + filial + local
        cursor.execute("""
            SELECT id FROM contagens
            WHERE sessao_id = ? AND codigo_item = ? AND filial = ? AND local = ? AND numero_contagem = 1
        """, sessao_id, codigo_item, filial, local)
        primeira = cursor.fetchone()
        pai_id = primeira.id if primeira else None

        # Recupera a(s) última(s) contagens somente para este sessao_id, codigo_item, filial e local com numero_contagem = ultimo_num_contagem
        cursor.execute("""
            SELECT id, saldo_sistema
            FROM contagens
            WHERE sessao_id = ? AND codigo_item = ? AND filial = ? AND local = ? AND numero_contagem = ?
        """, sessao_id, codigo_item, filial, local, ultimo_num_contagem)
        ultimas_contagens_local = cursor.fetchall()

        if not ultimas_contagens_local:
            flash('Não foram encontradas contagens da última rodada para este item neste local.', 'error')
            return redirect(url_for('visualizar_inventario', sessao_id=sessao_id))

        # Inserir apenas UMA nova recontagem por cada registro encontrado para este local (normalmente será só 1)
        for c_antiga in ultimas_contagens_local:
            # contagem_pai_id deve ser a primeira contagem (se existir), senão aponta para a contagem atual antiga
            pai_para_inserir = pai_id or c_antiga.id
            cursor.execute("""
                INSERT INTO contagens (sessao_id, usuario_id, codigo_item, filial, local, saldo_sistema, quantidade_contada, status_contagem, contagem_pai_id, numero_contagem)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            sessao_id, session['user_id'], codigo_item,
            filial, local, c_antiga.saldo_sistema,
            0, 'Pendente', pai_para_inserir, novo_num_contagem)

            # Atualiza o status da contagem antiga (apenas dessa local) para 'Recontagem Solicitada'
            cursor.execute("UPDATE contagens SET status_contagem = 'Recontagem Solicitada' WHERE id = ?", c_antiga.id)

        conn.commit()
        flash('Recontagem solicitada com sucesso para o local selecionado.', 'success')
    except Exception as e:
        conn.rollback()
        traceback.print_exc()
        flash(f'Erro ao solicitar recontagem: {e}', 'error')
    finally:
        if conn:
            conn.close()
    return redirect(url_for('visualizar_inventario', sessao_id=sessao_id_redirect))

@app.route('/item_history/<item_code>')
@login_required
def item_history(item_code):
    conn_inventario = get_db_connection(DB_INVENTARIO)
    cursor_inventario = conn_inventario.cursor()

    query = f"""
    SELECT 
        c.codigo_item, ISNULL(b1.B1_DESC, 'N/A') as B1_DESC,
        s.descricao as inventario_desc, c.data_contagem,
        c.saldo_sistema, c.quantidade_contada,
        (c.quantidade_contada - c.saldo_sistema) as diferenca,
        u.username
    FROM contagens c
    JOIN usuarios u ON c.usuario_id = u.id
    JOIN inventario_sessoes s ON c.sessao_id = s.id
    LEFT JOIN {DB_BI}.dbo.SB1010 b1 ON c.codigo_item COLLATE DATABASE_DEFAULT = b1.B1_COD COLLATE DATABASE_DEFAULT AND b1.D_E_L_E_T_ <> '*'
    WHERE c.codigo_item = ? AND s.D_E_L_E_T_ <> '*'
    ORDER BY c.data_contagem DESC
    """
    cursor_inventario.execute(query, item_code)
    historico = cursor_inventario.fetchall()
    conn_inventario.close()

    item_desc = historico[0].B1_DESC if historico else "Descrição não encontrada"

    return render_template('item_history.html', 
                           item_code=item_code, 
                           item_desc=item_desc, 
                           historico=historico, 
                           username=session.get('username'))

@app.route('/progress_dashboard')
@login_required
@role_required('Admin', 'Supervisor')
def selecionar_progress_dashboard():
    conn = get_db_connection(DB_INVENTARIO)
    cursor = conn.cursor()
    cursor.execute("SELECT id, descricao FROM inventario_sessoes WHERE status = 'Aberto' AND D_E_L_E_T_ <> '*' ORDER BY data_abertura DESC")
    inventarios_abertos = cursor.fetchall()
    conn.close()
    return render_template('selecionar_progress.html', inventarios=inventarios_abertos, username=session.get('username'))

@app.route('/progress_dashboard/<int:sessao_id>')
@login_required
@role_required('Admin', 'Supervisor')
def progress_dashboard(sessao_id):
    conn = get_db_connection(DB_INVENTARIO)
    cursor = conn.cursor()

    cursor.execute("SELECT id, descricao FROM inventario_sessoes WHERE id = ?", sessao_id)
    inventario = cursor.fetchone()
    if not inventario:
        flash("Inventário não encontrado.", "error")
        return redirect(url_for('selecionar_progress_dashboard'))

    # Query para buscar todos os dados necessários, incluindo custo
    query_detalhes = f"""
    SELECT 
        c.*, 
        (c.quantidade_contada - c.saldo_sistema) as diferenca,
        ISNULL(b1.B1_DESC, 'N/A') as B1_DESC,
        u.username,
        ISNULL(b2.B2_VATU1, 0) as custo_unitario
    FROM contagens c
    JOIN usuarios u ON c.usuario_id = u.id
    LEFT JOIN {DB_BI}.dbo.SB1010 b1 ON c.codigo_item COLLATE DATABASE_DEFAULT = b1.B1_COD COLLATE DATABASE_DEFAULT AND b1.D_E_L_E_T_ <> '*'
    LEFT JOIN {DB_BI}.dbo.SB2010 b2 ON c.codigo_item COLLATE DATABASE_DEFAULT = b2.B2_COD COLLATE DATABASE_DEFAULT
                                    AND c.filial COLLATE DATABASE_DEFAULT = b2.B2_FILIAL COLLATE DATABASE_DEFAULT
                                    AND c.local COLLATE DATABASE_DEFAULT = b2.B2_LOCAL COLLATE DATABASE_DEFAULT
                                    AND b2.D_E_L_E_T_ <> '*'
    WHERE c.sessao_id = ?
    """
    cursor.execute(query_detalhes, sessao_id)
    todas_as_contagens_raw = cursor.fetchall()

    # Lógica de correção de custo (idêntica à do dashboard analítico)
    custos_referencia = {}
    for c in todas_as_contagens_raw:
        custo_float = float(c.custo_unitario)
        if c.filial == '0201' and custo_float > 0:
            custos_referencia[c.codigo_item] = custo_float
    for c in todas_as_contagens_raw:
        custo_float = float(c.custo_unitario)
        if c.codigo_item not in custos_referencia and custo_float > 0:
             custos_referencia[c.codigo_item] = custo_float

    todas_as_contagens = []
    for c in todas_as_contagens_raw:
        contagem_dict = dict(zip([column[0] for column in c.cursor_description], c))
        contagem_dict['custo_unitario'] = float(contagem_dict['custo_unitario'])
        contagem_dict['saldo_sistema'] = float(contagem_dict['saldo_sistema'])
        contagem_dict['quantidade_contada'] = float(contagem_dict['quantidade_contada'])
        contagem_dict['diferenca'] = float(contagem_dict['diferenca'])
        if contagem_dict['filial'] == '0203' and contagem_dict['custo_unitario'] == 0:
            contagem_dict['custo_unitario'] = custos_referencia.get(contagem_dict['codigo_item'], 0.0)
        todas_as_contagens.append(contagem_dict)

    # Cálculo das métricas de progresso
    valor_total_contado = 0
    valor_sobra_parcial = 0
    valor_falta_parcial = 0

    for c in todas_as_contagens:
        custo = c['custo_unitario']
        valor_total_contado += c['quantidade_contada'] * custo
        diferenca_valor = c['diferenca'] * custo
        if diferenca_valor > 0:
            valor_sobra_parcial += diferenca_valor
        elif diferenca_valor < 0:
            valor_falta_parcial += abs(diferenca_valor)

    metricas = {
        'TotalContagens': len(todas_as_contagens),
        'ItensUnicosContados': len(set(c['codigo_item'] for c in todas_as_contagens)),
        'ValorTotalContado': valor_total_contado,
        'ValorSobraParcial': valor_sobra_parcial,
        'ValorFaltaParcial': valor_falta_parcial
    }
    
    # Prepara o feed com dados de custo
    todas_as_contagens.sort(key=lambda x: x['data_contagem'], reverse=True)
    feed = todas_as_contagens[:10]

    # Dados para o gráfico de usuários
    dados_usuarios = cursor.execute("SELECT u.username, COUNT(c.id) as NumContagens FROM contagens c JOIN usuarios u ON c.usuario_id = u.id WHERE c.sessao_id = ? GROUP BY u.username ORDER BY NumContagens DESC", sessao_id).fetchall()
    conn.close()
    
    chart_labels = [row.username for row in dados_usuarios]
    chart_data = [row.NumContagens for row in dados_usuarios]

    return render_template('progress_dashboard.html',
                           inventario=inventario,
                           metricas=metricas,
                           feed=feed,
                           chart_labels=chart_labels,
                           chart_data=chart_data,
                           username=session.get('username'))


# --- ROTAS DE GERAÇÃO DE ETIQUETAS ---
@app.route('/gerar_etiquetas', methods=['GET'])
@login_required
def gerar_etiquetas():
    produtos = None
    search_term = request.args.get('search_term', '')
    grupos_selecionados = request.args.getlist('grupos')
    filiais_selecionadas = request.args.getlist('filiais')
    locais_selecionados = request.args.getlist('locais')
    
    conn = get_db_connection(DB_BI)
    cursor = conn.cursor()
    
    cursor.execute("SELECT DISTINCT B2_FILIAL FROM SB2010 WHERE D_E_L_E_T_ <> '*' ORDER BY B2_FILIAL")
    filiais = [row.B2_FILIAL.strip() for row in cursor.fetchall()]
    cursor.execute("SELECT DISTINCT B2_LOCAL FROM SB2010 WHERE D_E_L_E_T_ <> '*' ORDER BY B2_LOCAL")
    locais = [row.B2_LOCAL.strip() for row in cursor.fetchall()]
    
    if request.args:
        # --- ALTERAÇÃO PRINCIPAL AQUI ---
        # A busca agora foca na SB1010 e junta a SB2010 apenas para filtros,
        # usando LEFT JOIN para não excluir produtos sem saldo/movimentação.
        query_parts = [
            "SELECT DISTINCT b1.B1_COD, b1.B1_DESC",
            "FROM SB1010 b1",
            "LEFT JOIN SB2010 b2 ON b1.B1_COD = b2.B2_COD AND b2.D_E_L_E_T_ <> '*'", # LEFT JOIN aqui
            "WHERE b1.D_E_L_E_T_ <> '*'" # Condição principal na SB1010
        ]
        params = []
        
        # O resto da lógica de filtros permanece a mesma
        if search_term:
            query_parts.append("AND (b1.B1_COD LIKE ? OR b1.B1_DESC LIKE ?)")
            like_term = f'%{search_term}%'
            params.extend([like_term, like_term])
            
        if grupos_selecionados:
            placeholders = ','.join(['?'] * len(grupos_selecionados))
            query_parts.append(f"AND SUBSTRING(b1.B1_COD, 1, 4) IN ({placeholders})")
            params.extend(grupos_selecionados)
            
        if filiais_selecionadas:
            placeholders = ','.join(['?'] * len(filiais_selecionadas))
            # O filtro de filial agora precisa verificar se b2 existe
            query_parts.append(f"AND b2.B2_FILIAL IN ({placeholders})")
            params.extend(filiais_selecionadas)
        
        if locais_selecionados:
            placeholders = ','.join(['?'] * len(locais_selecionados))
            # O filtro de local também precisa verificar se b2 existe
            query_parts.append(f"AND b2.B2_LOCAL IN ({placeholders})")
            params.extend(locais_selecionados)
            
        query_parts.append("ORDER BY b1.B1_COD")
        
        query = ' '.join(query_parts)
        cursor.execute(query, *params)
        produtos = cursor.fetchall()
        
    conn.close()
    
    return render_template('gerar_etiquetas.html', 
                           produtos=produtos, 
                           search_term=search_term, 
                           grupos_selecionados=grupos_selecionados,
                           filiais=filiais,
                           locais=locais,
                           filiais_selecionadas=filiais_selecionadas,
                           locais_selecionados=locais_selecionados,
                           username=session.get('username'))

@app.route('/gerar_etiquetas_massa_pdf', methods=['POST'])
@login_required
def gerar_etiquetas_massa_pdf():
    item_codes = request.form.getlist('item_codes')
    # Pega a escolha do utilizador (o padrão é 'qrcode' se nada for enviado)
    code_type = request.form.get('code_type', 'qrcode')

    if not item_codes:
        flash('Nenhum item selecionado para gerar etiquetas.', 'error')
        return redirect(url_for('gerar_etiquetas'))

    conn = get_db_connection(DB_BI)
    cursor = conn.cursor()
    
    placeholders = ','.join(['?'] * len(item_codes))
    query = f"SELECT B1_COD, B1_DESC FROM SB1010 WHERE B1_COD IN ({placeholders}) AND D_E_L_E_T_ <> '*'"
    cursor.execute(query, *item_codes)
    items = cursor.fetchall()
    conn.close()

    etiquetas_data = []
    for item in items:
        img_str = ""
        item_code_stripped = item.B1_COD.strip()
        
        # Buffer em memória para guardar a imagem
        buffered = BytesIO()

        # --- LÓGICA PARA ESCOLHER O TIPO DE CÓDIGO ---
        if code_type == 'barcode':
            # Gera um código de barras Code128 (bom para alfanuméricos)
            # A opção write_text=False evita que o código seja escrito debaixo da imagem
            barcode_img = Code128(item_code_stripped, writer=ImageWriter())
            barcode_img.write(buffered, options={'write_text': False, 'module_height': 10.0, 'font_size': 0, 'text_distance': 0})
        else: # O padrão é qrcode
            # Gera o QR Code
            qr = qrcode.QRCode(version=1, box_size=10, border=2)
            qr.add_data(item_code_stripped)
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")
            img.save(buffered, format="PNG")
        
        # Converte a imagem em memória para um formato que o HTML entende
        img_str = base64.b64encode(buffered.getvalue()).decode('utf-8')
        image_url = f"data:image/png;base64,{img_str}"
        etiquetas_data.append({'item': item, 'image_url': image_url})

    html_renderizado = render_template('etiquetas_massa_template.html', etiquetas=etiquetas_data)
    pdf = HTML(string=html_renderizado).write_pdf()

    return Response(pdf,
                    mimetype='application/pdf',
                    headers={'Content-Disposition': 'inline;filename=etiquetas.pdf'})

# --- Rotas de API ---
@app.route('/api/get_grupos')
@login_required
def get_grupos():
    try:
        conn = get_db_connection(DB_BI)
        cursor = conn.cursor()
        query = "SELECT DISTINCT SUBSTRING(B2_COD, 1, 4) as Grupo FROM SB2010 WHERE D_E_L_E_T_ <> '' ORDER BY Grupo"
        cursor.execute(query)
        grupos = [{'value': row.Grupo.strip(), 'text': row.Grupo.strip()} for row in cursor.fetchall()]
        conn.close()
        return jsonify(grupos)
    except Exception as e:
        print(f"Erro ao buscar grupos: {e}")
        return jsonify({'error': 'Não foi possível buscar a lista de grupos'}), 500

@app.route('/api/get_item_details', methods=['GET'])
@login_required
def get_item_details():
    """
    Busca os detalhes de um item, incluindo sua descrição, unidade de medida,
    e saldos em todas as filiais e locais de estoque para contagem.
    """
    user_id = session.get('user_id')
    conn_user = get_db_connection(DB_INVENTARIO)
    cursor_user = conn_user.cursor()
    cursor_user.execute("SELECT pode_ver_saldo FROM usuarios WHERE id = ?", user_id)
    user_permission = cursor_user.fetchone()
    conn_user.close()
    
    can_see_balance = user_permission.pode_ver_saldo if user_permission else False

    item_code = request.args.get('barcode', '')
    if not item_code:
        return jsonify({'error': 'Código do item não fornecido'}), 400
    
    conn = get_db_connection(DB_BI)
    conn_inventario = get_db_connection(DB_INVENTARIO)
    if not conn or not conn_inventario:
        return jsonify({'error': 'Não foi possível conectar ao banco de dados principal'}), 500
    
    cursor_bi = conn.cursor()
    cursor_inventario = conn_inventario.cursor()
    try:
        # Busca a descrição e UM do item na base de dados BI
        cursor_bi.execute("SELECT B1_DESC, B1_UM FROM SB1010 WHERE B1_COD = ? AND D_E_L_E_T_ <> '*'", item_code)
        item_info = cursor_bi.fetchone()
        
        if not item_info:
            return jsonify({'error': 'Item não encontrado ou inativo'}), 404

        description = item_info.B1_DESC.strip()
        unit_of_measure = item_info.B1_UM.strip()
        
        locations = []

        # Tenta buscar os saldos na base de dados BI
        query_saldos = """
            SELECT B2_FILIAL, B2_LOCAL, B2_QATU
            FROM SB2010
            WHERE B2_COD = ? AND D_E_L_E_T_ <> '*'
            ORDER BY B2_FILIAL, B2_LOCAL
        """
        cursor_bi.execute(query_saldos, item_code)
        saldos_existentes = cursor_bi.fetchall()

        if saldos_existentes:
            for row in saldos_existentes:
                filial = row.B2_FILIAL.strip()
                local = row.B2_LOCAL.strip()
                saldo_atual = float(row.B2_QATU or 0)

                # CORREÇÃO: Remove a condição D_E_L_E_T_ da consulta na SDB010
                cursor_bi.execute(
                    """
                    SELECT DB_LOCALIZ FROM SDB010 
                    WHERE DB_PRODUTO = ? AND DB_FILIAL = ? AND DB_LOCAL = ?
                    """,
                    item_code, filial, local
                )
                shelf_row = cursor_bi.fetchone()
                shelf_location = shelf_row.DB_LOCALIZ.strip() if shelf_row and shelf_row.DB_LOCALIZ else 'N/A'
                
                locations.append({
                    'saldo_atual': saldo_atual,
                    'filial': filial,
                    'local': local,
                    'shelf_location': shelf_location
                })
        else:
            # CORREÇÃO: Remove a condição D_E_L_E_T_ da consulta na contagens
            query_contagem_original = """
                SELECT DISTINCT filial, local FROM contagens 
                WHERE codigo_item = ?
            """
            cursor_inventario.execute(query_contagem_original, item_code)
            contagens_anteriores = cursor_inventario.fetchall()
            
            for row in contagens_anteriores:
                filial = row.filial.strip()
                local = row.local.strip()

                # CORREÇÃO: Remove a condição D_E_L_E_T_ da consulta na SDB010
                cursor_bi.execute(
                    """
                    SELECT DB_LOCALIZ FROM SDB010 
                    WHERE DB_PRODUTO = ? AND DB_FILIAL = ? AND DB_LOCAL = ?
                    """,
                    item_code, filial, local
                )
                shelf_row = cursor_bi.fetchone()
                shelf_location = shelf_row.DB_LOCALIZ.strip() if shelf_row and shelf_row.DB_LOCALIZ else 'N/A'
                
                locations.append({
                    'saldo_atual': 0.0,
                    'filial': filial,
                    'local': local,
                    'shelf_location': shelf_location
                })

        return jsonify({
            'description': description, 
            'unit_of_measure': unit_of_measure,
            'locations': locations,
            'can_see_balance': can_see_balance
        })

    except pyodbc.Error as e:
        print(f"Erro no banco de dados: {e}")
        return jsonify({'error': 'Erro ao consultar o banco de dados'}), 500
    finally:
        conn.close()
        conn_inventario.close()

@app.route('/api/save_count', methods=['POST'])
@login_required
def save_count():
    data = request.json
    counts_data = data.get('counts')
    sessao_id = data.get('sessao_id')
    user_id = session.get('user_id')

    if not all([counts_data, sessao_id]):
        return jsonify({'error': 'Dados de contagem ou ID do inventário em falta'}), 400

    conn = get_db_connection(DB_INVENTARIO)
    cursor = conn.cursor()
    
    try:
        for count in counts_data:
            codigo_item = count.get('codigo_item')
            filial = count.get('filial')
            local = count.get('local')
            quantidade_contada = count.get('quantidade_contada')

            cursor.execute("""
                SELECT id FROM contagens 
                WHERE sessao_id = ? AND codigo_item = ? AND filial = ? AND local = ? 
                AND status_contagem = 'Pendente'
            """, sessao_id, codigo_item, filial, local)
            contagem_pendente = cursor.fetchone()

            if contagem_pendente:
                cursor.execute("""
                    UPDATE contagens 
                    SET quantidade_contada = ?, usuario_id = ?, data_contagem = GETDATE(), status_contagem = 'OK'
                    WHERE id = ?
                """, quantidade_contada, user_id, contagem_pendente.id)
            else:
                if quantidade_contada is not None:
                    cursor.execute(
                        """
                        INSERT INTO contagens (codigo_item, filial, local, quantidade_contada, saldo_sistema, usuario_id, sessao_id, numero_contagem, status_contagem) 
                        VALUES (?, ?, ?, ?, ?, ?, ?, 1, 'OK')
                        """,
                        codigo_item, filial, local,
                        quantidade_contada, count.get('saldo_sistema'),
                        user_id, sessao_id
                    )

        conn.commit()
        return jsonify({'success': f'{len(counts_data)} contagem(ns) salva(s) com sucesso!'})
    except Exception as e:
        conn.rollback()
        traceback.print_exc()
        return jsonify({'error': f'Erro ao salvar contagem: {e}'}), 500
    finally:
        conn.close()

@app.route('/api/update_count/<int:contagem_id>', methods=['PUT'])
@login_required
def update_count(contagem_id):
    data = request.json
    nova_quantidade = data.get('quantidade')

    if nova_quantidade is None:
        return jsonify({'error': 'Nova quantidade não fornecida'}), 400

    conn = get_db_connection(DB_INVENTARIO)
    cursor = conn.cursor()
    try:
        cursor.execute("UPDATE contagens SET quantidade_contada = ?, status_contagem = 'OK' WHERE id = ?", nova_quantidade, contagem_id)
        conn.commit()
        if cursor.rowcount == 0:
            return jsonify({'error': 'Contagem não encontrada'}), 404
        return jsonify({'success': 'Contagem atualizada com sucesso!'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': f'Erro ao atualizar: {e}'}), 500
    finally:
        conn.close()

@app.route('/api/delete_count/<int:contagem_id>', methods=['DELETE'])
@login_required
@role_required('Admin', 'Supervisor')
def delete_count(contagem_id):
    conn = get_db_connection(DB_INVENTARIO)
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM contagens WHERE id = ?", contagem_id)
        conn.commit()
        if cursor.rowcount == 0:
            return jsonify({'error': 'Contagem não encontrada'}), 404
        return jsonify({'success': 'Contagem excluída com sucesso!'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': f'Erro ao excluir: {e}'}), 500
    finally:
        conn.close()

@app.route('/api/ask_chatbot', methods=['POST'])
@login_required
def ask_chatbot():
    user_question = request.json.get('question')
    if not user_question:
        return jsonify({'answer': 'Por favor, faça uma pergunta.'})

    prompt = f"""
    Sua tarefa é converter uma pergunta em linguagem natural para uma consulta T-SQL segura e de apenas leitura.
    Você tem acesso a dois bancos de dados: '{DB_INVENTARIO}' e '{DB_BI}'.

    Estrutura do banco de dados '{DB_INVENTARIO}':
    - usuarios(id, username, filial, role, D_E_L_E_T_)
    - inventario_sessoes(id, descricao, status, data_abertura, data_fechamento, D_E_L_E_T_)
    - contagens(id, codigo_item, filial, local, quantidade_contada, saldo_sistema, sessao_id, usuario_id, status_contagem)

    Estrutura do banco de dados '{DB_BI}':
    - SB1010(B1_COD, B1_DESC, D_E_L_E_T_) -> Descrição dos produtos
    - SB2010(B2_COD, B2_FILIAL, B2_LOCAL, B2_QATU, B2_DMOV, D_E_L_E_T_) -> Saldo e movimentação

    Regras:
    1. Gere APENAS a consulta T-SQL. Sem explicações ou texto adicional.
    2. Use APENAS consultas SELECT. NUNCA use UPDATE, DELETE, INSERT, etc.
    3. Sempre inclua a condição `D_E_L_E_T_ <> '*'` nas tabelas que a possuem.
    4. Para cruzar dados entre os bancos, use o nome completo, ex: `{DB_BI}.dbo.SB1010`.
    5. A data de hoje é {datetime.now().strftime('%Y-%m-%d')}.

    Exemplos:
    - Pergunta: "quantas contagens o admin fez?"
    - SQL: SELECT COUNT(*) FROM contagens c JOIN usuarios u ON c.usuario_id = u.id WHERE u.username = 'admin' AND u.D_E_L_E_T_ <> '*';
    - Pergunta: "qual a descrição do item 1234?"
    - SQL: SELECT B1_DESC FROM {DB_BI}.dbo.SB1010 WHERE B1_COD = '1234' AND D_E_L_E_T_ <> '*';
    - Pergunta: "quais inventários estão abertos?"
    - SQL: SELECT descricao FROM inventario_sessoes WHERE status = 'Aberto' AND D_E_L_E_T_ <> '*';

    Pergunta do usuário: "{user_question}"
    """
    
    try:
        chatHistory = [{"role": "user", "parts": [{"text": prompt}]}]
        payload = {"contents": chatHistory}
        api_key = "AIzaSyDwThzfEAf25l1efhMP_YtmYw49a0L_rWA" # IMPORTANTE: Insira sua chave aqui
        api_url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash-preview-05-20:generateContent?key={api_key}"
        
        response = requests.post(api_url, json=payload, headers={'Content-Type': 'application/json'})
        response.raise_for_status()
        result = response.json()
        
        sql_query_raw = result['candidates'][0]['content']['parts'][0]['text']
        sql_query = sql_query_raw.replace('```sql', '').replace('```', '').strip()
        select_pos = sql_query.lower().find('select')
        if select_pos != -1:
            sql_query = sql_query[select_pos:]

        if not sql_query.lower().startswith('select'):
            print(f"Resposta da IA não era um SELECT válido: {sql_query_raw}")
            return jsonify({'answer': 'Desculpe, só posso executar consultas de leitura (SELECT).'})

        conn = get_db_connection(DB_INVENTARIO)
        cursor = conn.cursor()
        cursor.execute(sql_query)
        
        rows = cursor.fetchall()
        if not rows:
            return jsonify({'answer': 'A consulta não retornou resultados.'})

        columns = [column[0] for column in cursor.description]
        answer_html = '<table class="min-w-full divide-y divide-gray-200"><thead><tr>'
        for col in columns:
            answer_html += f'<th class="px-2 py-1 bg-gray-50 text-left text-xs font-medium text-gray-500 uppercase">{col}</th>'
        answer_html += '</tr></thead><tbody class="bg-white divide-y divide-gray-200">'
        for row in rows:
            answer_html += '<tr>'
            for value in row:
                answer_html += f'<td class="px-2 py-1 text-sm text-gray-700">{value}</td>'
            answer_html += '</tr>'
        answer_html += '</tbody></table>'
        
        conn.close()
        return jsonify({'answer': answer_html})

    except requests.exceptions.RequestException as e:
        return jsonify({'answer': f'Erro ao comunicar com a IA: {e}'})
    except pyodbc.Error as e:
        return jsonify({'answer': f'Erro ao consultar o banco de dados: A IA pode ter gerado uma consulta inválida.'})
    except Exception as e:
        return jsonify({'answer': f'Ocorreu um erro inesperado: {e}'})
    

@app.route('/gerar_etiquetas_pdf_zebra', methods=['POST'])
@login_required
def gerar_etiquetas_pdf_zebra():
    item_codes = request.form.getlist('item_codes')
    code_type = request.form.get('code_type', 'qrcode')

    if not item_codes:
        flash('Nenhum item selecionado para gerar as etiquetas.', 'error')
        return redirect(url_for('gerar_etiquetas'))

    conn = get_db_connection(DB_BI)
    cursor = conn.cursor()
    placeholders = ','.join(['?'] * len(item_codes))
    query = f"SELECT B1_COD, B1_DESC FROM SB1010 WHERE B1_COD IN ({placeholders}) AND D_E_L_E_T_ <> '*'"
    cursor.execute(query, *item_codes)
    items = cursor.fetchall()
    conn.close()

    etiquetas_data = []
    for item in items:
        item_code_stripped = item.B1_COD.strip()
        buffered = BytesIO()

        if code_type == 'barcode':
            barcode_img = Code128(item_code_stripped, writer=ImageWriter())
            barcode_img.write(buffered, options={'write_text': False, 'module_height': 10.0})
        else: # qrcode
            qr = qrcode.QRCode(version=1, box_size=10, border=2)
            qr.add_data(item_code_stripped)
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")
            img.save(buffered, format="PNG")
        
        img_str = base64.b64encode(buffered.getvalue()).decode('utf-8')
        image_url = f"data:image/png;base64,{img_str}"
        etiquetas_data.append({'item': item, 'image_url': image_url})

    # Renderiza o novo template específico para a Zebra
    html_renderizado = render_template('etiqueta_zebra_template.html', etiquetas=etiquetas_data)
    pdf = HTML(string=html_renderizado).write_pdf()

    return Response(pdf,
                    mimetype='application/pdf',
                    headers={'Content-Disposition': 'inline;filename=etiquetas_zebra.pdf'})

@app.route('/api/save_consolidated_count', methods=['POST'])
@login_required
def save_consolidated_count():
    data = request.json
    sessao_id = data.get('sessao_id')
    user_id = session.get('user_id')
    item_code = data.get('item_code')
    group_name = data.get('group_name')
    total_quantity_str = data.get('total_quantity')

    if not all([sessao_id, user_id, item_code, group_name, total_quantity_str]):
        return jsonify({'error': 'Dados insuficientes para processar a contagem.'}), 400

    try:
        total_quantity = float(total_quantity_str)
        filiais_do_grupo = ALMOXARIFADOS_GRUPO.get(group_name)
        if not filiais_do_grupo:
            return jsonify({'error': f'Grupo de almoxarifado "{group_name}" não encontrado.'}), 404

        conn_inv = get_db_connection(DB_INVENTARIO)
        cursor_inv = conn_inv.cursor()

        placeholders = ','.join(['?'] * len(filiais_do_grupo))
        cursor_inv.execute(f"""
            SELECT id, filial, local, saldo_sistema FROM contagens 
            WHERE sessao_id = ? AND codigo_item = ? AND status_contagem = 'Pendente' AND filial IN ({placeholders})
        """, sessao_id, item_code, *filiais_do_grupo)
        contagens_pendentes = cursor_inv.fetchall()

        if contagens_pendentes:
            saldos_locais = [{'filial': c.filial, 'local': c.local, 'saldo': float(c.saldo_sistema)} for c in contagens_pendentes]
        else:
            conn_bi = get_db_connection(DB_BI)
            cursor_bi = conn_bi.cursor()
            saldos_locais = []
            for filial in filiais_do_grupo:
                cursor_bi.execute(
                    "SELECT B2_LOCAL, B2_QATU FROM SB2010 WHERE B2_COD = ? AND B2_FILIAL = ? AND D_E_L_E_T_ <> '*' ORDER BY B2_LOCAL",
                    item_code, filial
                )
                for row in cursor_bi.fetchall():
                    saldos_locais.append({'filial': filial, 'local': row.B2_LOCAL.strip(), 'saldo': float(row.B2_QATU or 0.0)})
            conn_bi.close()
            
            if not saldos_locais:
                cursor_inv.execute(f"SELECT DISTINCT filial, local FROM contagens WHERE codigo_item = ? AND filial IN ({placeholders})", item_code, *filiais_do_grupo)
                for row in cursor_inv.fetchall():
                    saldos_locais.append({'filial': row.filial.strip(), 'local': row.local.strip(), 'saldo': 0.0})

        total_saldo_sistema = sum(item['saldo'] for item in saldos_locais)
        alocacoes_finais = { (item['filial'], item['local']): 0.0 for item in saldos_locais }

        if total_quantity >= total_saldo_sistema:
            surplus = total_quantity - total_saldo_sistema
            for local_info in saldos_locais:
                alocacoes_finais[(local_info['filial'], local_info['local'])] = local_info['saldo']
            
            target_for_surplus = next(( (li['filial'], li['local']) for li in sorted(saldos_locais, key=lambda x: x['saldo'], reverse=True) if li['saldo'] > 0.0), None)
            if not target_for_surplus and saldos_locais:
                target_for_surplus = (saldos_locais[0]['filial'], saldos_locais[0]['local'])
            if target_for_surplus:
                alocacoes_finais[target_for_surplus] += surplus
        else:
            remaining_quantity = total_quantity
            saldos_locais.sort(key=lambda x: x['saldo'], reverse=True)
            for local_info in saldos_locais:
                allocated = min(local_info['saldo'], remaining_quantity)
                alocacoes_finais[(local_info['filial'], local_info['local'])] = allocated
                remaining_quantity -= allocated
                if remaining_quantity <= 0: break

        if contagens_pendentes:
            for pendente in contagens_pendentes:
                quantidade_alocada = alocacoes_finais.get((pendente.filial, pendente.local), 0.0)
                cursor_inv.execute("""
                    UPDATE contagens SET quantidade_contada = ?, usuario_id = ?, data_contagem = GETDATE(), status_contagem = 'OK'
                    WHERE id = ?
                """, quantidade_alocada, user_id, pendente.id)
        else:
            cursor_inv.execute(f"DELETE FROM contagens WHERE sessao_id = ? AND codigo_item = ? AND filial IN ({placeholders})", sessao_id, item_code, *filiais_do_grupo)
            for local_info in saldos_locais:
                quantidade_alocada = alocacoes_finais.get((local_info['filial'], local_info['local']), 0.0)
                cursor_inv.execute(
                    """
                    INSERT INTO contagens (sessao_id, usuario_id, codigo_item, filial, local, saldo_sistema, quantidade_contada, status_contagem, numero_contagem)
                    VALUES (?, ?, ?, ?, ?, ?, ?, 'OK', 1)
                    """,
                    sessao_id, user_id, item_code, local_info['filial'], local_info['local'], local_info['saldo'], quantidade_alocada
                )
        
        conn_inv.commit()
        conn_inv.close()
        return jsonify({'success': f'Contagem de {total_quantity} unidades alocada com sucesso.'})

    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': f'Ocorreu um erro interno: {e}'}), 500
    
@app.route('/contagem_consolidada')
@login_required
def contagem_consolidada_page():
    # ... (código existente para buscar sessões e grupos)
    conn = get_db_connection(DB_INVENTARIO)
    cursor = conn.cursor()
    cursor.execute("SELECT id, descricao FROM inventario_sessoes WHERE status = 'Aberto' AND D_E_L_E_T_ <> '*' ORDER BY data_abertura DESC")
    sessoes_abertas = cursor.fetchall()
    conn.close()
    
    grupos_almoxarifado = list(ALMOXARIFADOS_GRUPO.keys())
    
    return render_template('contagem_consolidada.html', 
                           sessoes=sessoes_abertas, 
                           grupos=grupos_almoxarifado,
                           ALMOXARIFADOS_GRUPO=ALMOXARIFADOS_GRUPO,
                           username=session.get('username'))


@app.route('/analise_estoque_parado', methods=['GET', 'POST'])
@login_required
@role_required('Admin', 'Supervisor')
def analise_estoque_parado():
    itens = None
    valor_total = 0
    data_selecionada = (datetime.now() - timedelta(days=180)).strftime('%Y-%m-%d')
    data_selecionada_formatada = ''

    filtros = {
        'grupos': request.form.getlist('grupos'),
        'filiais': request.form.getlist('filiais'),
        'locais': request.form.getlist('locais'),
        'search_term': request.form.get('search_term', '')
    }

    conn_bi = get_db_connection(DB_BI)
    cursor_bi = conn_bi.cursor()
    
    cursor_bi.execute("SELECT DISTINCT B2_FILIAL FROM SB2010 WHERE D_E_L_E_T_ <> '*' ORDER BY B2_FILIAL")
    lista_filiais = [row.B2_FILIAL.strip() for row in cursor_bi.fetchall()]
    cursor_bi.execute("SELECT DISTINCT B2_LOCAL FROM SB2010 WHERE D_E_L_E_T_ <> '*' ORDER BY B2_LOCAL")
    lista_locais = [row.B2_LOCAL.strip() for row in cursor_bi.fetchall()]

    if request.method == 'POST':
        data_corte_str = request.form.get('data_corte')
        data_selecionada = data_corte_str
        
        data_corte_db = datetime.strptime(data_corte_str, '%Y-%m-%d').strftime('%Y%m%d')
        data_selecionada_formatada = datetime.strptime(data_corte_str, '%Y-%m-%d').strftime('%d/%m/%Y')

        query_base = """
        SELECT b2.B2_COD, MAX(b1.B1_DESC) as B1_DESC, MAX(b2.B2_DMOV) as UltimaMov,
               SUM(b2.B2_QATU) as SaldoTotal, MAX(ISNULL(b2.B2_CMFIM1, 0)) as CustoUnitario
        FROM SB2010 b2
        LEFT JOIN SB1010 b1 ON b2.B2_COD = b1.B1_COD AND b1.D_E_L_E_T_ <> '*'
        WHERE b2.B2_QATU > 0 AND b2.D_E_L_E_T_ <> '*'
        """
        params = []
        
        if filtros['search_term']:
            query_base += " AND (b2.B2_COD LIKE ? OR b1.B1_DESC LIKE ?)"
            like_term = f"%{filtros['search_term']}%"
            params.extend([like_term, like_term])
            
        if filtros['grupos']:
            placeholders = ','.join(['?'] * len(filtros['grupos']))
            query_base += f" AND SUBSTRING(b2.B2_COD, 1, 4) IN ({placeholders})"
            params.extend(filtros['grupos'])

        if filtros['filiais']:
            placeholders = ','.join(['?'] * len(filtros['filiais']))
            query_base += f" AND b2.B2_FILIAL IN ({placeholders})"
            params.extend(filtros['filiais'])

        if filtros['locais']:
            placeholders = ','.join(['?'] * len(filtros['locais']))
            query_base += f" AND b2.B2_LOCAL IN ({placeholders})"
            params.extend(filtros['locais'])

        query_final = f"""
        {query_base}
        GROUP BY b2.B2_COD
        HAVING MAX(b2.B2_DMOV) <= ?
        ORDER BY UltimaMov ASC
        """
        params.append(data_corte_db)
        
        cursor_bi.execute(query_final, *params)
        
        resultados_brutos = cursor_bi.fetchall()
        itens_processados = []
        for row in resultados_brutos:
            custo = float(row.CustoUnitario or 0)
            saldo = float(row.SaldoTotal or 0)
            valor_total_item = saldo * custo
            
            # --- CORREÇÃO APLICADA AQUI ---
            # Verifica se o campo UltimaMov não está vazio antes de o converter
            ultima_mov_formatada = 'N/A' # Define um valor padrão
            if row.UltimaMov and row.UltimaMov.strip():
                ultima_mov_formatada = datetime.strptime(row.UltimaMov.strip(), '%Y%m%d').strftime('%d/%m/%Y')
            
            itens_processados.append({
                'codigo': row.B2_COD.strip(), 'descricao': row.B1_DESC.strip(),
                'ultima_mov': ultima_mov_formatada, # Usa a variável formatada e segura
                'saldo': saldo, 'custo': custo, 'valor_total': valor_total_item
            })
            valor_total += valor_total_item
        
        itens = itens_processados

    conn_bi.close()

    return render_template('estoque_parado.html',
                           itens=itens, valor_total=valor_total,
                           data_selecionada=data_selecionada,
                           data_selecionada_formatada=data_selecionada_formatada,
                           username=session.get('username'),
                           lista_filiais=lista_filiais, lista_locais=lista_locais,
                           filtros_selecionados=filtros)

# --- Execução da Aplicação ---
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5002, debug=True)